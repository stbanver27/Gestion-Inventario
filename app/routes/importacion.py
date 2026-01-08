from fastapi import APIRouter, UploadFile, File, HTTPException
from app.utils import load_json, save_json, DATA_DIR
from openpyxl import load_workbook

router = APIRouter(prefix="/productos", tags=["Productos"])

PRODUCTOS_FILE = DATA_DIR / "productos.json"

def next_id(items: list[dict]) -> int:
    return (max((i.get("id", 0) for i in items), default=0) + 1)

@router.post("/importar_excel")
async def importar_productos_excel(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sube un archivo .xlsx")

    content = await file.read()

    try:
        wb = load_workbook(filename=bytes(content))
    except Exception:
        # openpyxl no soporta "filename=bytes" directo en todas las versiones
        # alternativa: usar BytesIO
        from io import BytesIO
        try:
            wb = load_workbook(BytesIO(content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel inválido: {e}")

    sheet_name = "productos"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="La hoja debe llamarse 'productos'")

    ws = wb[sheet_name]

    # Leer headers (fila 1)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    required = {"nombre", "categoria", "precio"}
    if not required.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias. Requiere: {sorted(list(required))}"
        )

    # Map header -> index
    idx = {h: i for i, h in enumerate(headers)}

    productos = load_json(PRODUCTOS_FILE, default=[])
    prod_by_id = {p["id"]: p for p in productos}

    creados = 0
    actualizados = 0
    errores = []

    # Procesar filas desde la 2
    for row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c+1).value for c in range(len(headers))]

        # Saltar filas vacías
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def get(col, default=None):
            j = idx.get(col)
            if j is None:
                return default
            v = row[j]
            return default if v is None else v

        try:
            pid = get("id", None)
            pid = int(pid) if pid not in (None, "") else None

            nombre = str(get("nombre", "")).strip()
            categoria = str(get("categoria", "")).strip()
            precio = float(get("precio", 0) or 0)
            costo = float(get("costo", 0) or 0)
            stock = int(float(get("stock", 0) or 0))

            if not nombre or not categoria:
                raise ValueError("nombre/categoria vacíos")
            if precio < 0 or costo < 0 or stock < 0:
                raise ValueError("precio/costo/stock no pueden ser negativos")

            if pid and pid in prod_by_id:
                # actualizar
                prod = prod_by_id[pid]
                prod["nombre"] = nombre
                prod["categoria"] = categoria
                prod["precio"] = precio
                prod["costo"] = costo
                prod["stock"] = stock
                actualizados += 1
            else:
                # crear
                new_id = pid if pid and pid not in prod_by_id else next_id(productos)
                prod = {
                    "id": new_id,
                    "nombre": nombre,
                    "categoria": categoria,
                    "precio": precio,
                    "costo": costo,
                    "stock": stock,
                }
                productos.append(prod)
                prod_by_id[new_id] = prod
                creados += 1

        except Exception as e:
            errores.append({"fila": row_num, "error": str(e)})

    save_json(PRODUCTOS_FILE, productos)

    return {
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores[:50],  # evita respuestas gigantes
        "total_errores": len(errores),
    }
